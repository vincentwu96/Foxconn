import PyPDF2, io
import urllib.request as urllib2
import openpyxl
import os
from xlrd import open_workbook,cellname
import time
# Line 2603 in pdf.py from PyPDF2 should be modified to: text +=  "\n" + _text

start = time.time()
print("FA Ticket Parser written by Vincent Wu\nLoading please wait...")

# Format and combine all pages to FA_Tickets_All
all_book = open_workbook('FA_Tickets_Parse.xls')
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'] = "Serial No"
sheet['B1'] = "GUTI"
sheet['C1'] = "FAID"
sheet['D1'] = "FA Status"
sheet['E1'] = "Family"
sheet['F1'] = "LOGOP"
sheet['G1'] = "Status"
sheet['H1'] = "Release ID"
sheet['I1'] = "Start"
sheet['J1'] = "Complete"
sheet['K1'] = "Test Set"
sheet['L1'] = "Test Case"
sheet['M1'] = "Code"
sheet['N1'] = "Type Code"
sheet['O1'] = "Fault Code"
sheet['P1'] = "Subfault Code"
sheet['Q1'] = "Message"
sheet['R1'] = "Notes"
sheet['S1'] = "FA Result"
row = 2
for y in range(0, all_book.nsheets):    # Could just add another for loop
    read_sheet = all_book.sheet_by_index(y)
    for x in range(1,read_sheet.nrows):
        sheet['A' + str(row)] = read_sheet.cell_value(x,0)
        sheet['B' + str(row)] = read_sheet.cell_value(x,1)
        sheet['C' + str(row)] = read_sheet.cell_value(x,2)
        sheet['D' + str(row)] = read_sheet.cell_value(x,3)
        sheet['E' + str(row)] = read_sheet.cell_value(x,4)
        sheet['F' + str(row)] = read_sheet.cell_value(x,5)
        sheet['G' + str(row)] = read_sheet.cell_value(x,6)
        sheet['H' + str(row)] = read_sheet.cell_value(x,7)
        sheet['I' + str(row)] = read_sheet.cell_value(x,8)
        sheet['J' + str(row)] = read_sheet.cell_value(x,9)
        sheet['K' + str(row)] = read_sheet.cell_value(x,10)
        sheet['L' + str(row)] = read_sheet.cell_value(x,11)
        sheet['M' + str(row)] = read_sheet.cell_value(x,12)
        sheet['N' + str(row)] = read_sheet.cell_value(x,13)
        sheet['O' + str(row)] = read_sheet.cell_value(x,14)
        sheet['P' + str(row)] = read_sheet.cell_value(x,15)
        sheet['Q' + str(row)] = read_sheet.cell_value(x,16)
        sheet['R' + str(row)] = read_sheet.cell_value(x,17)
        sheet['S' + str(row)] = read_sheet.cell_value(x,18)
        int(row)
        row +=1    
wb.save('FA_Tickets_All.xls')

book = open_workbook('FA_Tickets_All.xls')
read_sheet = book.sheet_by_index(0)
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
 
row = 1
for x in range(0,read_sheet.nrows):
    sheet['A' + str(row)] = read_sheet.cell_value(x,0)  # Serial
    sheet['B' + str(row)] = read_sheet.cell_value(x,3)  # FA Status
    sheet['C' + str(row)] = read_sheet.cell_value(x,16).rstrip() # Message
    sheet['D' + str(row)] = read_sheet.cell_value(x,18).rstrip() # FA Result
    int(row)
    row +=1
sheet['F1'] = "Original Partnum"
sheet['G1'] = "Original Serialnum"
sheet['H1'] = "Original Vendor"
wb.save('FA_Tickets.xlsx')

row = 1
FAID = 0
for x in range(0,read_sheet.nrows):
    if(not x == 0):
        print("Parsing", read_sheet.cell_value(x,0))
    URL = "http://10.250.250.141:7021/TDMSREPORTRUNNER/JasperServlet?REPORT=FASUMMARY&FAID="
    if(read_sheet.col_values(2)[FAID] == "FAID"):       # FA Actions
        sheet['E' + str(row)] = "FA Actions"
    elif(read_sheet.col_values(2)[FAID] == "-"):        # No FAID
        sheet['E' + str(row)] = "-"
    elif(read_sheet.col_values(3)[FAID] == "OPEN"):     # OPEN status
        sheet['E' + str(row)] = "-"
    elif(sheet['A' + str(row)] == None or sheet['A' + str(row)] == ""):     # Blank lines in FA_Tickets_Parse.xls
        sheet['E' + str(row)] = ""
    elif(read_sheet.col_values(3)[FAID] == None or read_sheet.col_values(3)[FAID] == ""):   # Blank lines in FA_Tickets_Parse.xls
        sheet['E' + str(row)] = ""
    else:
        URL += read_sheet.col_values(2)[FAID]
        
        # Parsing PDF
        req = urllib2.Request(URL, headers={'User-Agent' : "Magic Browser"})
        remote_file = urllib2.urlopen(req).read()
        memory_file = io.BytesIO(remote_file)

        read_pdf = PyPDF2.PdfFileReader(memory_file)
        number_of_pages = read_pdf.getNumPages()

        text = ""
        for i in range(0, number_of_pages):
            pageObj = read_pdf.getPage(i)
            page = pageObj.extractText()
            text += page

        # Formatting PDF Output
        format1 = text.split("kok.yee@foxconn.com", 1)[1]   # Delete front elements
        format1 = format1.rsplit("Description", 1)[0]       # Delete back elements
        format2 = format1.split("\n")
        output = ""                                         # Description and Date
        Partnum = ""
        Serialnum = ""
        Vendor = ""
        counter = 1        # Counter for Description and Date
        counter2 = 1       # Counter for Partnum, Serialnum, Vendor
        flag = False
        for x in range(0,len(format2)):
            if(format2[x] == "Vendor"):
                Partnum += str(counter2) + ") " + format2[x+1] + "\n"
                Serialnum += str(counter2) + ") " + format2[x+2] + "\n"
                Vendor += str(counter2) + ") " + format2[x+3] + "\n"
                int(counter2)
                counter2+=1
            if(format2[x] == "kok.yee@foxconn.com"):
                if(flag == True):
                    break
                y = x
                z = y+4
                output+= str(counter) + ") " + format2[x+1] + " "
                int(counter)
                counter+=1
                for y in range(y,len(format2)):
                    if(format2[y] == "Description"):
                        break
                    elif(format2[y] == "Vendor"):
                        z+=11
                    elif(format2[y] == "INITIAL"):
                        last = len(format2) - 1 - format2[::-1].index('kok.yee@foxconn.com')
                        output = ' '.join(output.split(' ')[:-1])
                        output = ' '.join(output.split(' ')[:-1])
                        output += format2[last+1] + " "
                        for i in range(last+5, len(format2)):
                            output += format2[i] + "\n"
                        flag = True
                        break
                    else:
                        if(y>z):
                            output += format2[y] + "\n"
                            
        output = output.rstrip()
        # Last catch for "ORACLE" (No Description)
        if "ORACLE" in output:
            output = "-"
        else:
            sheet['E' + str(row)] = output
        Partnum = Partnum.rstrip()
        Serialnum = Serialnum.rstrip()
        Vendor = Vendor.rstrip()
        sheet['F' + str(row)] = Partnum
        sheet['G' + str(row)] = Serialnum
        sheet['H' + str(row)] = Vendor
    int(row)
    row+=1
    FAID+=1
    wb.save('FA_Tickets.xlsx')



# WMS Report
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')

sheet['A1'] = 'Test'
sheet['B1'] = 'Family'
sheet['C1'] = 'SN'
sheet['D1'] = 'MSG'
sheet['E1'] = 'FA Conclusion'
sheet['F1'] = 'FA Actions'
sheet['G1'] = 'Category'
sheet['H1'] = 'Code'
sheet['I1'] = 'Type'
sheet['J1'] = 'Category for Metrix Report'

row = 2
all_book = open_workbook('FA_Tickets_Parse.xls')
read_sheet = all_book.sheet_by_index(0)
for y in range(0, all_book.nsheets):
    read_sheet = all_book.sheet_by_index(y)
    for x in range(1,read_sheet.nrows):
        sheet['A' + str(row)] = read_sheet.cell_value(x,5)
        sheet['B' + str(row)] = read_sheet.cell_value(x,4)
        sheet['C' + str(row)] = read_sheet.cell_value(x,0)
        sheet['D' + str(row)] = read_sheet.cell_value(x,16)
        sheet['E' + str(row)] = read_sheet.cell_value(x,18)
        # FA Actions in other file
        sheet['G' + str(row)] = read_sheet.cell_value(x,12)
        sheet['H' + str(row)] = read_sheet.cell_value(x,13)
        sheet['I' + str(row)] = read_sheet.cell_value(x,14)
        int(row)
        row+=1

FA_Actions = openpyxl.load_workbook('FA_Tickets.xlsx')
FA_Actions_Sheet = FA_Actions.get_sheet_by_name('Sheet')

row = 2
for x in range(1,FA_Actions_Sheet.max_row):
    sheet['F' + str(row)] = FA_Actions_Sheet['E' + str(row)].value
    int(row)
    row+=1

# Failure Analysis by Category
row = 2
for x in range(1,FA_Actions_Sheet.max_row): 
    if(sheet['E' + str(row)].value == '-'):
        sheet['J' + str(row)] = "Open"
    if(not sheet['E' + str(row)].value == None and not sheet['G' + str(row)].value == None):    # Check for None
        # If reseat/swap in FA Conclusion and Category is COMPONENT
        if("reseat" in sheet['E' + str(row)].value.lower() and sheet['G' + str(row)].value == "COMPONENT"):
            sheet['J' + str(row)] = "5) Connectivity"
        if("swap" in sheet['E' + str(row)].value.lower() and sheet['G' + str(row)].value == "COMPONENT"):
            sheet['J' + str(row)] = "5) Connectivity"
        # If replace in FA Conclusion
        if("replace" in sheet['E' + str(row)].value.lower()):
            sheet['J' + str(row)] = "4) Component"
    # If Type is SWDLOAD
    if(not sheet['I' + str(row)].value == None):
        if(sheet['I' + str(row)].value =="SWDLOAD"):
            sheet['J' + str(row)] = "1) Test Process (SWDL)"
    # If Category = TESTPROCESS and CODE = PERSONNEL
    if(not sheet['G' + str(row)].value == None and not sheet['H' + str(row)].value == None):
        if(sheet['G' + str(row)].value == "TESTPROCESS" and sheet['H' + str(row)].value == "PERSONNEL"):
            sheet['J' + str(row)] = "7) Assembly TMS"
    if(not sheet['E' + str(row)].value == None and not sheet['G' + str(row)].value == None):
        if("correct" in sheet['E' + str(row)].value.lower() and sheet['G' + str(row)].value == "ASSYPROCESS"):
            sheet['J' + str(row)] = "7) Assembly WMS"
    int(row)
    row+=1    
wb.save("WMS Report.xlsx")

end = time.time()
time_elapsed = end - start
print("Completed in ", time_elapsed, "s", sep='')
os.system("pause")

