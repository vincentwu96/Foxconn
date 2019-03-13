import openpyxl
import os
import time
from xlrd import open_workbook,cellname

start = time.time()
print("WMS Report Generator written by Vincent Wu\nLoading please wait...")

# WMS Report
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')

sheet['A1'] = 'Test'
sheet['B1'] = 'Family'
sheet['C1'] = 'SN'
sheet['D1'] = 'MSG'
sheet['E1'] = 'FA Conclusion'
sheet['F1'] = 'FA Actions'
sheet['G1'] = 'Category'
sheet['H1'] = 'Code'
sheet['I1'] = 'Type'
sheet['J1'] = 'Category for Metrix Report'

row = 2
all_book = open_workbook('FA_Tickets_Parse.xls')
read_sheet = all_book.sheet_by_index(0)
for y in range(0, all_book.nsheets):
    read_sheet = all_book.sheet_by_index(y)
    for x in range(1,read_sheet.nrows):
        sheet['A' + str(row)] = read_sheet.cell_value(x,5)
        sheet['B' + str(row)] = read_sheet.cell_value(x,4)
        sheet['C' + str(row)] = read_sheet.cell_value(x,0)
        sheet['D' + str(row)] = read_sheet.cell_value(x,16)
        sheet['E' + str(row)] = read_sheet.cell_value(x,18)
        # FA Actions in other file
        sheet['G' + str(row)] = read_sheet.cell_value(x,12)
        sheet['H' + str(row)] = read_sheet.cell_value(x,13)
        sheet['I' + str(row)] = read_sheet.cell_value(x,14)
        int(row)
        row+=1

FA_Actions = openpyxl.load_workbook('FA_Tickets.xlsx')
FA_Actions_Sheet = FA_Actions.get_sheet_by_name('Sheet')


row = 2
for x in range(1,FA_Actions_Sheet.max_row):
    sheet['F' + str(row)] = FA_Actions_Sheet['E' + str(row)].value
    int(row)
    row+=1


# Failure Analysis by Category
row = 2
for x in range(1,FA_Actions_Sheet.max_row): 
    if(sheet['E' + str(row)].value == '-'):
        sheet['J' + str(row)] = "Open"
    if(not sheet['E' + str(row)].value == None and not sheet['G' + str(row)].value == None):    # Check for None
        # If reseat/swap in FA Conclusion and Category is COMPONENT
        if("reseat" in sheet['E' + str(row)].value.lower() and sheet['G' + str(row)].value == "COMPONENT"):
            sheet['J' + str(row)] = "5) Connectivity"
        if("swap" in sheet['E' + str(row)].value.lower() and sheet['G' + str(row)].value == "COMPONENT"):
            sheet['J' + str(row)] = "5) Connectivity"
        # If replace in FA Conclusion
        if("replace" in sheet['E' + str(row)].value.lower()):
            sheet['J' + str(row)] = "4) Component"
    # If Type is SWDLOAD
    if(not sheet['I' + str(row)].value == None):
        if(sheet['I' + str(row)].value =="SWDLOAD"):
            sheet['J' + str(row)] = "1) Test Process (SWDL)"
    # If Category = TESTPROCESS and CODE = PERSONNEL
    if(not sheet['G' + str(row)].value == None and not sheet['H' + str(row)].value == None):
        if(sheet['G' + str(row)].value == "TESTPROCESS" and sheet['H' + str(row)].value == "PERSONNEL"):
            sheet['J' + str(row)] = "7) Assembly TMS"
    if(not sheet['E' + str(row)].value == None and not sheet['G' + str(row)].value == None):
        if("correct" in sheet['E' + str(row)].value.lower() and sheet['G' + str(row)].value == "ASSYPROCESS"):
            sheet['J' + str(row)] = "7) Assembly WMS"
    int(row)
    row+=1    


wb.save("WMS Report.xlsx")
print("Completed!")

